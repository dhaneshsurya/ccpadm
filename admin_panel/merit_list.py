"""Merit list helpers — class/program-wise ranking by education percentage."""

from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO

from admissions.models import StudentAdmission
from admissions.services import build_education_list
from courses.constants import PROGRAM_LEVEL_DISPLAY
from courses.utils import get_program_level_for_name

MERIT_DEFAULT_STATUSES = ('Submitted', 'Approved', 'Pending')

STATUS_FILTER_CHOICES = (
    ('applied', 'Submitted / Approved / Pending'),
    ('submitted', 'Submitted only'),
    ('approved', 'Approved only'),
    ('pending', 'Pending only'),
    ('all_active', 'All except Draft & Rejected'),
)

MERIT_EXPORT_HEADERS = (
    'Program', 'Program Level', 'Merit Rank', 'Application No', 'Reg No',
    'Student Name', "Father's Name", 'Status', 'Sort Basis',
    'Merit %', '12th %', 'Graduation %', 'Mobile', 'Email',
)


@dataclass
class MeritEntry:
    admission: StudentAdmission
    program_level: str
    merit_percentage: float | None
    merit_basis: str
    twelfth_percentage: float | None
    graduation_percentage: float | None


def parse_percentage(value) -> float | None:
    """Parse percentage strings such as 85, 85.5, or 85.5%."""
    if value is None:
        return None
    text = str(value).strip().replace('%', '')
    if not text or text in ('-', '—'):
        return None
    text = text.replace(',', '.')
    match = re.search(r'(\d+(?:\.\d+)?)', text)
    if not match:
        return None
    try:
        return round(float(match.group(1)), 2)
    except ValueError:
        return None


def _percentage_from_education_rows(admission, class_keywords: tuple[str, ...]) -> float | None:
    for row in build_education_list(admission):
        class_name = (
            row.get('className')
            or row.get('ClassName')
            or row.get('Class')
            or ''
        ).lower()
        if not any(keyword in class_name for keyword in class_keywords):
            continue
        pct = parse_percentage(row.get('percentage') or row.get('Percentage'))
        if pct is not None:
            return pct
    return None


def get_twelfth_percentage(admission) -> float | None:
    pct = parse_percentage(getattr(admission, 'percentage12', None))
    if pct is not None:
        return pct
    return _percentage_from_education_rows(
        admission,
        ('12', 'xii', 'inter', 'higher secondary', 'h.s.c', 'hsc'),
    )


def get_graduation_percentage(admission) -> float | None:
    pct = parse_percentage(getattr(admission, 'percentage_grad', None))
    if pct is not None:
        return pct
    return _percentage_from_education_rows(
        admission,
        ('grad', 'bachelor', 'degree', 'b.a', 'b.sc', 'b.com', 'ba ', 'bsc', 'bcom'),
    )


def resolve_merit_basis(program_level: str) -> str:
    if program_level == 'PG':
        return 'Graduation'
    return '12th'


def build_merit_entry(admission: StudentAdmission) -> MeritEntry:
    program_type = (admission.program_type or '').strip()
    program_level = get_program_level_for_name(program_type)
    twelfth_pct = get_twelfth_percentage(admission)
    grad_pct = get_graduation_percentage(admission)

    if program_level == 'PG':
        merit_pct = grad_pct
        merit_basis = 'Graduation'
    else:
        merit_pct = twelfth_pct
        merit_basis = '12th'

    return MeritEntry(
        admission=admission,
        program_level=program_level or '—',
        merit_percentage=merit_pct,
        merit_basis=merit_basis,
        twelfth_percentage=twelfth_pct,
        graduation_percentage=grad_pct,
    )


def _statuses_for_filter(status_filter: str) -> tuple[str, ...]:
    mapping = {
        'applied': MERIT_DEFAULT_STATUSES,
        'submitted': ('Submitted',),
        'approved': ('Approved',),
        'pending': ('Pending',),
        'all_active': ('Submitted', 'Approved', 'Pending', 'Draft'),
    }
    return mapping.get(status_filter, MERIT_DEFAULT_STATUSES)


def _merit_sort_key(entry: MeritEntry):
    pct = entry.merit_percentage
    name = (entry.admission.full_name or '').lower()
    app_no = entry.admission.application_no or ''
    if pct is None:
        return (1, 0, name, app_no)
    return (0, -pct, name, app_no)


def get_merit_program_choices(status_filter: str = 'applied') -> list[str]:
    statuses = _statuses_for_filter(status_filter)
    names = (
        StudentAdmission.objects.filter(status__in=statuses)
        .exclude(program_type='')
        .values_list('program_type', flat=True)
        .distinct()
    )
    return sorted({(name or '').strip() for name in names if (name or '').strip()}, key=str.lower)


def build_merit_list_groups(
    program_filter: str = 'ALL',
    status_filter: str = 'applied',
) -> list[dict]:
    statuses = _statuses_for_filter(status_filter)
    admissions = (
        StudentAdmission.objects.filter(status__in=statuses)
        .exclude(program_type='')
        .order_by('program_type', '-submitted_date', '-created_date')
    )
    if program_filter and program_filter != 'ALL':
        admissions = admissions.filter(program_type=program_filter)

    grouped: dict[str, list[MeritEntry]] = {}
    for admission in admissions:
        program = (admission.program_type or '').strip() or 'Not Assigned'
        grouped.setdefault(program, []).append(build_merit_entry(admission))

    result = []
    for program in sorted(grouped.keys(), key=str.lower):
        entries = sorted(grouped[program], key=_merit_sort_key)
        program_level = entries[0].program_level if entries else ''
        ranked = []
        for rank, entry in enumerate(entries, start=1):
            ranked.append({
                'rank': rank,
                'entry': entry,
                'admission': entry.admission,
            })
        result.append({
            'program': program,
            'program_level': program_level,
            'program_level_display': PROGRAM_LEVEL_DISPLAY.get(program_level, program_level or '—'),
            'merit_basis': resolve_merit_basis(program_level),
            'count': len(ranked),
            'rows': ranked,
        })
    return result


def get_status_filter_label(status_filter: str) -> str:
    for key, label in STATUS_FILTER_CHOICES:
        if key == status_filter:
            return label
    return STATUS_FILTER_CHOICES[0][1]


def iter_merit_export_rows(groups: list[dict]):
    """Yield flat rows for CSV / Excel export."""
    for group in groups:
        for row in group['rows']:
            admission = row['admission']
            entry = row['entry']
            yield [
                group['program'],
                group['program_level_display'],
                row['rank'],
                admission.application_no or '',
                admission.reg_no or '',
                admission.full_name or '',
                admission.father_name or '',
                admission.status or '',
                entry.merit_basis,
                entry.merit_percentage if entry.merit_percentage is not None else '',
                entry.twelfth_percentage if entry.twelfth_percentage is not None else '',
                entry.graduation_percentage if entry.graduation_percentage is not None else '',
                admission.mobile or '',
                admission.email or '',
            ]


def build_merit_list_workbook(
    groups: list[dict],
    program_filter: str = 'ALL',
    status_filter: str = 'applied',
) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Merit List'

    title_font = Font(name='Arial', bold=True, size=14)
    meta_font = Font(name='Arial', size=10, color='444444')
    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E40AF')
    body_font = Font(name='Arial', size=10)
    pct_cols = {10, 11, 12}

    program_label = program_filter if program_filter != 'ALL' else 'All Programs'
    status_label = get_status_filter_label(status_filter)
    last_col = get_column_letter(len(MERIT_EXPORT_HEADERS))

    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = 'Merit List — Program Wise'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'] = f'Program: {program_label}  |  Status: {status_label}'
    ws['A2'].font = meta_font
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    header_row = 4
    for col_idx, header in enumerate(MERIT_EXPORT_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    data_row = header_row + 1
    for values in iter_merit_export_rows(groups):
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_idx, value=value)
            cell.font = body_font
            if col_idx in pct_cols and value not in ('', None):
                cell.number_format = '0.00'
        data_row += 1

    ws.freeze_panes = f'A{header_row + 1}'
    ws.auto_filter.ref = f'A{header_row}:{last_col}{max(header_row, data_row - 1)}'

    for col_idx in range(1, len(MERIT_EXPORT_HEADERS) + 1):
        letter = get_column_letter(col_idx)
        max_len = len(str(MERIT_EXPORT_HEADERS[col_idx - 1]))
        for row_idx in range(header_row, data_row):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def merit_export_filename(program_filter: str = 'ALL', extension: str = 'xlsx') -> str:
    if program_filter and program_filter != 'ALL':
        safe = re.sub(r'[^\w\-]+', '_', program_filter).strip('_')
        return f'merit_list_{safe}.{extension}'
    return f'merit_list.{extension}'